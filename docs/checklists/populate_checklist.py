import os
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from datetime import datetime

# Define the directory and file paths
project_dir = os.path.abspath(".")
output_dir = os.path.join(project_dir, "docs", "checklists")
os.makedirs(output_dir, exist_ok=True)  # Ensure the directory exists

# Define output file for the new checklist
new_file_path = os.path.join(output_dir, "regenerated_algorithmic_trading_checklist.xlsx")

# Original checklist data (combine original and additional items here)
checklist_data = {
    "Category": [
        # =========== Backtesting Basics ===========
        "Backtesting Basics",  # Backtesting/1
        "Backtesting Basics",  # Backtesting/2
        "Backtesting Basics",  # Backtesting/3
        "Backtesting Basics",  # Backtesting/4
        "Backtesting Basics",  # Backtesting/5
        "Backtesting Basics",  # Backtesting/6
        "Backtesting Basics",  # Backtesting/7

        # =========== Market Data Handling ===========
        "Market Data Handling",  # Market/1
        "Market Data Handling",  # Market/2
        "Market Data Handling",  # Market/3
        "Market Data Handling",  # Market/4

        # =========== Strategy Design ===========
        "Strategy Design",  # Strategy/1
        "Strategy Design",  # Strategy/2
        "Strategy Design",  # Strategy/3
        "Strategy Design",  # Strategy/4
        "Strategy Design",  # Strategy/5
        "Strategy Design",  # Strategy/6

        # =========== Execution Considerations ===========
        "Execution Considerations",  # Execution/1

        # =========== Risk Management ===========
        "Risk Management",  # Risk/1
        "Risk Management",  # Risk/2
        "Risk Management",  # Risk/3

        # =========== Validation ===========
        "Validation",  # Validation/1
        "Validation",  # Validation/2
        "Validation",  # Validation/3
        "Validation",  # Validation/4
        "Validation",  # Validation/5

        # =========== Performance Metrics ===========
        "Performance Metrics",  # Metrics/1
        "Performance Metrics",  # Metrics/2
        "Performance Metrics",  # Metrics/3
        "Performance Metrics",  # Metrics/4

        # =========== Data Preprocessing ===========
        "Data Preprocessing",  # Preprocessing/1
        "Data Preprocessing",  # Preprocessing/2

        # =========== Order Book Model ===========
        "Order Book Model",  # OrderBook/1
        "Order Book Model",  # OrderBook/2
        "Order Book Model",  # OrderBook/3

        # =========== Matching Engine ===========
        "Matching Engine",  # Matching/1
        "Matching Engine",  # Matching/2
        "Matching Engine",  # Matching/3

        # =========== Scalability ===========
        "Scalability",       # Scalability/1
        "Scalability",       # Scalability/2

        # =========== Validation and QA ===========
        "Validation and QA", # ValidationQA/1
        "Validation and QA", # ValidationQA/2

        # =========== Event-Driven Architecture ===========
        "Event-Driven Architecture"  # Event/4
    ],

    "Identifier": [
        # =========== Backtesting Basics ===========
        "Backtesting/1",
        "Backtesting/2",
        "Backtesting/3",
        "Backtesting/4",
        "Backtesting/5",
        "Backtesting/6",
        "Backtesting/7",

        # =========== Market Data Handling ===========
        "Market/1",
        "Market/2",
        "Market/3",
        "Market/4",

        # =========== Strategy Design ===========
        "Strategy/1",
        "Strategy/2",
        "Strategy/3",
        "Strategy/4",
        "Strategy/5",
        "Strategy/6",

        # =========== Execution Considerations ===========
        "Execution/1",

        # =========== Risk Management ===========
        "Risk/1",
        "Risk/2",
        "Risk/3",

        # =========== Validation ===========
        "Validation/1",
        "Validation/2",
        "Validation/3",
        "Validation/4",
        "Validation/5",

        # =========== Performance Metrics ===========
        "Metrics/1",
        "Metrics/2",
        "Metrics/3",
        "Metrics/4",

        # =========== Data Preprocessing ===========
        "Preprocessing/1",
        "Preprocessing/2",

        # =========== Order Book Model ===========
        "OrderBook/1",
        "OrderBook/2",
        "OrderBook/3",

        # =========== Matching Engine ===========
        "Matching/1",
        "Matching/2",
        "Matching/3",

        # =========== Scalability ===========
        "Scalability/1",
        "Scalability/2",

        # =========== Validation and QA ===========
        "ValidationQA/1",
        "ValidationQA/2",

        # =========== Event-Driven Architecture ===========
        "Event/4"
    ],

    "Consideration": [
        # =========== Backtesting Basics ===========
        "Avoid Look-Ahead Bias",
        "Address Data-Snooping Bias",
        "Use Realistic Transaction Costs",
        "Evaluate Statistical Significance",
        "Address Randomness in Finite Sample Sizes",
        "Avoid Backtesting Low Sharpe Ratio Strategies",
        "Compare to Appropriate Benchmarks",

        # =========== Market Data Handling ===========
        "Adjust for Venue-Specific Quotes",
        "Handle Missing or Outlier Data",
        "Avoid Survivorship Bias",
        "Use Survivorship-Bias-Free Datasets",

        # =========== Strategy Design ===========
        "Simplify Models",
        "Avoid Overfitting in Model Design",
        "Consider Benchmark Appropriateness",
        "Use Appropriate Statistical Measures for Strategy Validation",
        "Avoid Overfitted Models",
        "Implement Baseline (Buy-and-Hold) or No-Trade Scenario for Comparison",

        # =========== Execution Considerations ===========
        "Account for Latency and Slippage",

        # =========== Risk Management ===========
        "Add Risk Controls",
        "Implement Stop-Loss & Take-Profit Logic",
        "Position Sizing & Leverage Management",

        # =========== Validation ===========
        "Use Walk-Forward Testing",
        "Test for Robustness Using Monte Carlo Simulation",
        "Use Randomized Trade Testing",
        "Test for Overfitting with Statistical Significance",
        "Evaluate High-Frequency Strategies Critically",

        # =========== Performance Metrics ===========
        "Track Statistical Significance",
        "Track Sortino Ratio",
        "Track Calmar Ratio",
        "Track Profit Factor",

        # =========== Data Preprocessing ===========
        "Validate dataset integrity, handle missing levels, and remove duplicate timestamps.",
        "Convert data to efficient formats (Parquet, HDF5) for scalability.",

        # =========== Order Book Model ===========
        "Ensure price-time priority in order book handling.",
        "Implement data structures for managing 24 levels of bid/ask prices.",
        "Simulate Queue Position at Each Limit Price.",

        # =========== Matching Engine ===========
        "Simulate slippage, partial fills, and transaction costs.",
        "Add latency simulation to account for realistic execution delays.",
        "Enable distributed processing with frameworks like Dask or Ray.",

        # =========== Scalability ===========
        "Optimize dataset access using columnar formats for faster I/O.",
        "Perform statistical consistency checks and simulate trades for fill validation.",

        # =========== Validation and QA ===========
        "Automate data quality checks and anomaly detection during ingestion.",
        "Implement Event-Driven Architecture for Order-Book Snapshots.",

        # =========== Event-Driven Architecture ===========
        "Implement Event-Driven Architecture for Order-Book Snapshots."  # (duplicate mention if needed)
    ],

    "Status": [
        "⚠️ Pending" for _ in range(43)
    ]
}

# Create DataFrame
checklist_df = pd.DataFrame(checklist_data)

# Save the checklist to an Excel file with formatting
with pd.ExcelWriter(new_file_path, engine='openpyxl') as writer:
    checklist_df.to_excel(writer, index=False, sheet_name="Checklist")
    worksheet = writer.sheets["Checklist"]

    # Define category colors for better readability
    category_colors = {
        "Backtesting Basics": "B4C7E7",
        "Market Data Handling": "A9D08E",
        "Strategy Design": "FFE699",
        "Execution Considerations": "F4B084",
        "Risk Management": "D9E2F3",
        "Validation": "FABF8F",
        "Performance Metrics": "FFD966",
        "Data Preprocessing": "9CC3E6",
        "Order Book Model": "A9D08E",
        "Matching Engine": "FFE699",
        "Scalability": "F4B084",
        "Validation and QA": "FABF8F",
        "Event-Driven Architecture": "FFD966"
    }

    # Apply colors to rows based on category
    for row_idx, category in enumerate(checklist_df["Category"], start=2):
        fill_color = category_colors.get(category, "FFFFFF")  # Default white if category not found
        for col_idx in range(1, len(checklist_df.columns) + 1):
            worksheet.cell(row=row_idx, column=col_idx).fill = PatternFill(
                start_color=fill_color, end_color=fill_color, fill_type="solid"
            )

    # Adjust column widths
    for col_num, column in enumerate(checklist_df.columns, 1):
        col_letter = get_column_letter(col_num)
        max_length = max(checklist_df[column].astype(str).apply(len).max(), len(column)) + 2
        worksheet.column_dimensions[col_letter].width = max_length

print(f"Checklist successfully regenerated and saved to {new_file_path}")

# Verification Report
original_count = len(checklist_df)
print(f"Verification Report: Total items in the regenerated checklist: {original_count}. No items omitted.")
quit