import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill

# Checklist data: This dictionary contains all the relevant information for the checklist.
checklist_data = {
    "Category": [
        "Event-Driven Architecture", "Event-Driven Architecture", "Event-Driven Architecture",
        "Data Handling", "Data Handling", "Data Handling",
        "Order Matching Engine", "Order Matching Engine",
        "Scalability", "Scalability",
        "Validation and Statistical Significance", "Validation and Statistical Significance"
    ],
    "Identifier": [
        "Event/1", "Event/2", "Event/3",
        "Data/1", "Data/2", "Data/3",
        "Order/1", "Order/2",
        "Scale/1", "Scale/2",
        "Validation/1", "Validation/2"
    ],
    "Consideration": [
        "Adopt Complex Event Processing (CEP)",
        "Use Event-Driven Programming for Instantaneous Reaction",
        "Minimize Latency in Event-Triggered Processes",
        "Implement Efficient Storage for Level 2 Data",
        "Ensure Real-Time Data Processing Capabilities",
        "Preprocess Data to Remove Anomalies and Handle Gaps",
        "Develop a Market Microstructure Simulation",
        "Incorporate Execution Dynamics (e.g., Slippage, Latency, Partial Fills)",
        "Use Parallel or Distributed Computing Frameworks",
        "Optimize for Tick-Level Scalability and Large Datasets",
        "Perform Monte Carlo Simulations to Test Strategy Robustness",
        "Apply Walk-Forward Testing and Cross-Validation"
    ],
    "Description/Action": [
        "Leverage CEP frameworks to handle tick-by-tick updates and complex multi-condition rules.",
        "Ensure the system reacts instantly to market events like new ticks or news updates, avoiding polling delays.",
        "Optimize event processing pipelines to minimize latency and ensure timely responses for high-frequency strategies.",
        "Use columnar storage formats (e.g., Parquet, HDF5) or in-memory databases (e.g., Redis) for fast Level 2 data access.",
        "Integrate tools for real-time data handling (e.g., Kafka, Flink) to process incoming streams efficiently.",
        "Clean data to address gaps, errors, and inconsistencies before backtesting or live trading.",
        "Simulate market order books, replicating price-time priority and other matching rules to ensure accurate backtests.",
        "Incorporate models for slippage, latency, and partial fills to mimic real-world execution dynamics.",
        "Leverage parallel computing libraries (e.g., Dask, Ray) or distributed frameworks (e.g., Spark) to handle computationally intensive workloads.",
        "Design the system to scale efficiently for large datasets and tick-level granularity across multiple symbols.",
        "Run Monte Carlo simulations to evaluate how strategies perform across thousands of hypothetical scenarios, ensuring robustness.",
        "Use walk-forward testing and cross-validation to confirm the strategy generalizes to unseen data."
    ],
    "Status": [
        "⚠️ Pending", "⚠️ Pending", "⚠️ Pending",
        "⚠️ Pending", "⚠️ Pending", "⚠️ Pending",
        "⚠️ Pending", "⚠️ Pending",
        "⚠️ Pending", "⚠️ Pending",
        "⚠️ Pending", "⚠️ Pending"
    ]
}

# Define file paths: Specify the directory and file name for saving the checklist.
output_dir = "./docs/checklists"  # Directory to save the file
output_file = "advanced_backtesting_checklist.xlsx"  # File name for the checklist
os.makedirs(output_dir, exist_ok=True)  # Create the directory if it doesn't exist
file_path = os.path.join(output_dir, output_file)  # Full path to the file

# Create DataFrame: Convert the checklist data into a pandas DataFrame for easier processing.
checklist_df = pd.DataFrame(checklist_data)

# Save to Excel with auto-adjusted column widths and color-coding
with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
    # Write the DataFrame to an Excel file
    checklist_df.to_excel(writer, index=False, sheet_name='Checklist')
    worksheet = writer.sheets['Checklist']  # Access the worksheet for further formatting
    
    # Define category colors: Assign colors to each category for better visualization.
    category_colors = {
        "Event-Driven Architecture": "B4C7E7",  # Light Blue
        "Data Handling": "A9D08E",  # Light Green
        "Order Matching Engine": "FFE699",  # Light Yellow
        "Scalability": "F4B084",  # Light Orange
        "Validation and Statistical Significance": "FABF8F"  # Peach
    }

    # Apply colors: Fill each row with its category-specific color.
    for row_idx, category in enumerate(checklist_df['Category'], start=2):
        fill_color = category_colors.get(category, "FFFFFF")  # Default to white if category not found
        for col_idx in range(1, len(checklist_df.columns) + 1):
            worksheet.cell(row=row_idx, column=col_idx).fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

    # Adjust column widths: Automatically size columns based on their content for better readability.
    for col_num, column in enumerate(checklist_df.columns, 1):
        col_letter = get_column_letter(col_num)  # Get the column letter (e.g., A, B, C)
        max_length = max(checklist_df[column].astype(str).apply(len).max(), len(column)) + 2  # Calculate max content width
        worksheet.column_dimensions[col_letter].width = max_length  # Set column width

# Notify user: Indicate that the file has been successfully created and saved.
print(f"Checklist successfully updated with enhanced readability and saved to {file_path}")
